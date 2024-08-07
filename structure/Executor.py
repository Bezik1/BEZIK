import pyautogui
import time
import openpyxl
import subprocess
import sys
from os import path, environ, startfile
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import requests
from bs4 import BeautifulSoup
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.options import Options as EdgeOptions

from utils.excel_formats import map_format

PROGRAM_EXTENSIONS = {
    "python": "py",
    "javascript": "js",
    "typescript": "ts",
    "css": "css",
    "html": "html",
    "java": "java",
    "php": "php",
    "txt": "txt",
    "csv": "csv"
}

WEBDRIVER_SCRIPTS = {
    "highlight": """
    arguments[0].style.border='3px solid lightblue';
    arguments[0].style.backgroundColor='lightblue';
    """,
    "scroll": "arguments[0].scrollIntoView({ behavior: 'smooth', block: 'center' });",
}

START_PROGRAMS = {
    "python": "python",
    "txt": "notepad",
}

EXCEL_DICT = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

class Executor():
    def __init__(self) -> None:
        self.cache = {}

    def text_analize(self, text):
        pass

    def command_list_preparation(self, text):
        return text.split("\n")

    def join_quotes(self, lst):
        result = []
        i = 0
        while i < len(lst):
            if lst[i] == "'":
                quoted_str = lst[i]
                i += 1
                while i < len(lst) and lst[i] != "'":
                    quoted_str += lst[i]
                    i += 1
                if i < len(lst) and lst[i] == "'":
                    quoted_str += lst[i]
                result.append(quoted_str)
            else:
                result.append(lst[i])
            i += 1
        return result

    def merge_between_parentheses(self, lst):
        result = []
        temp = []
        inside_parentheses = False
        
        for item in lst:
            if item == '[':
                inside_parentheses = True
                temp = []
            elif item == ']':
                inside_parentheses = False
                result.append(''.join(temp))
            elif inside_parentheses:
                temp.append(item)
            else:
                result.append(item)
        
        return result

    def execute(self, command_list):
        command_type = command_list[0]
        match command_type:
            case 'webdriver':
                command_list = self.join_quotes(command_list)

                mode = command_list[1]

                edge_options = EdgeOptions()
                edge_options.use_chromium = True
                edge_options.add_argument("--start-maximized")
                edge_options.add_argument("--disable-popup-blocking")
                edge_options.add_argument("--disable-notifications")

                driver = webdriver.Edge(options=edge_options)
                
                match mode:
                    case "search":
                        page_source = driver.page_source
                        search_text_param = command_list[3][1:-1]

                        url = command_list[2]
                        driver.get(url)

                        try:
                            element = driver.find_element(By.XPATH, f"//*[contains(text(), '{search_text_param}')]")

                            driver.execute_script(WEBDRIVER_SCRIPTS["highlight"], element)
                            driver.execute_script(WEBDRIVER_SCRIPTS["scroll"], element)
                        except:
                            print("Nie odnaleziono tekstu")
                    case "find":
                        search_text_param = " ".join(command_list[2:])
                        search_text_param = search_text_param[0].capitalize() + search_text_param[1:]

                        url = f'https://www.google.com/search?q={search_text_param}'
                        driver.get(url)

                        try:
                            accept_all_button_xpath = "//button[@id='L2AGLb']"
                            accept_all_button = WebDriverWait(driver, 10).until(
                                EC.element_to_be_clickable((By.XPATH, accept_all_button_xpath))
                            )
                            accept_all_button.click()
                            elements = driver.find_elements(By.XPATH, f"//*[contains(text(), '{search_text_param}')]")

                            links = []
                            if elements:
                                for element in elements:
                                    if element.tag_name.lower() == 'a':
                                        link = element.get_attribute('href')
                                        if link:
                                            links.append(link)

                                    driver.execute_script(WEBDRIVER_SCRIPTS["highlight"], element)
                            
                            information_div = driver.find_element(By.CLASS_NAME, "kno-rdesc")

                            self.cache["informations"] = information_div.text[5:]
                            self.cache["links"] = links
                        except:
                            print("Nie odnaleziono tekstu")
                        
                    case _:
                        print("Nieprawidłowy tryb")
                driver.quit()

            case 'close':
                subprocess.run(["taskkill", "/F", "/IM", f'{" ".join(command_list[1:])[0:]}.exe'])

            case 'open':
                command_str = " ".join(command_list[1:])

                pyautogui.press('winleft')

                pyautogui.write(command_str)
                pyautogui.press('enter')

                return command_str
            case 'excel':                
                command_list = self.merge_between_parentheses(command_list)
                operation = command_list[1]

                match operation:
                    case "columns":
                        wb = openpyxl.Workbook()
                        sheet = wb.active
                
                        column_names = command_list[2].split(',')
                        column_values = []

                        data_list = command_list[3:]
                        
                        for str_list in data_list:
                            column_values.append(str_list.split(','))
                        
                        for i, name in enumerate(column_names):
                            sheet[f'{EXCEL_DICT[i]}1'] = name

                        for i in range(len(column_names)):
                            for j in range(len(column_values[0])):
                                sheet[f'{EXCEL_DICT[i]}{j+2}'] = column_values[i][j]

                        file_path = 'example.xlsx'
                        wb.save(file_path)
                        startfile(file_path)
                    case "load":
                        wb = openpyxl.Workbook()
                        sheet = wb.active

                        desktop_path = path.join(path.join(environ['USERPROFILE']), 'Desktop')
                        data_file_path = f'{desktop_path}/{command_list[2]}.txt'
                        file = open(data_file_path, "r")

                        for row_idx, line in enumerate(file.readlines(), start=1):
                            cells = line.strip().split()
                            for col_idx, cell_value in enumerate(cells, start=1):
                                cell = wb.active.cell(row=row_idx, column=col_idx, value=cell_value)
                                cell.number_format = map_format(cell_value)
                        
                        file_path = 'example.xlsx'
                        wb.save(file_path)
                        startfile(file_path)
                    case "add":
                        desktop_path = path.join(path.join(environ['USERPROFILE']), 'Desktop')
                        data_file_path = f'{desktop_path}/{command_list[2]}.xlsx'

                        wb = openpyxl.load_workbook(data_file_path)
                        sheet = wb.active

                        column_names = command_list[3].split(',')
                        column_values = []

                        data_list = command_list[4:]
                        
                        for str_list in data_list:
                            column_values.append(str_list.split(','))

                        def get_empty_column():
                            for i in range(20):
                                if sheet.cell(row=1, column=i+1).value is None:
                                    return i
                            return -1
                        
                        empty_col_idx = get_empty_column()
                        
                        for i, name in enumerate(column_names):
                            sheet[f'{EXCEL_DICT[i+empty_col_idx]}1'] = name
                        
                        for i in range(len(column_names)):
                            for j in range(len(column_values[0])):
                                cell = sheet[f'{EXCEL_DICT[i + empty_col_idx]}{j + 2}']
                                cell.value = column_values[i][j]
                                cell.number_format = map_format(column_values[i][j])

                        wb.save(data_file_path)
                        startfile(data_file_path)
                    case _:
                        print("Nieprawidłowa komenda excela")
            
            case 'write':
                extension = PROGRAM_EXTENSIONS[command_list[1]]
                program = f'''{" ".join(command_list[2:])[1:-1].replace("true", "True").replace("false", "False")}'''[0:]

                desktop_path = path.join(path.join(environ['USERPROFILE']), 'OneDrive\Pulpit')
                desktop_file_path = path.join(desktop_path, f'file.{extension}')
                file_path = f'file.{extension}'

                writer_file = open(file_path, "w")
                writer_file.write(program)
                writer_file.close()

                desktop_writer_file = open(desktop_file_path, "w")
                desktop_writer_file.write(program)
                desktop_writer_file.close()

                file = open(file_path, "a")
                subprocess.call([START_PROGRAMS[command_list[1]], file_path],creationflags=subprocess.CREATE_NEW_CONSOLE)
                
            case _:
                print("Nieprawidłowa komenda")
        return self.cache